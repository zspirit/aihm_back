import csv
import io
import secrets
from datetime import datetime, timezone
from uuid import UUID

import structlog
from celery import shared_task
from openpyxl import load_workbook

logger = structlog.get_logger()


def get_sync_session():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session

    from app.core.config import get_settings

    settings = get_settings()
    sync_url = settings.DATABASE_URL.replace("+asyncpg", "")
    engine = create_engine(sync_url)
    return Session(engine)


@shared_task(name="bulk_import.process", bind=True, max_retries=1)
def process_csv_import(self, bulk_import_id: str):
    logger.info("bulk_import_start", import_id=bulk_import_id)

    session = get_sync_session()
    try:
        from app.models.bulk_import import BulkImport
        from app.models.candidate import Candidate
        from app.models.consent import Consent
        from app.services.storage import download_file

        bulk_import = session.get(BulkImport, UUID(bulk_import_id))
        if not bulk_import:
            logger.error("bulk_import_not_found", import_id=bulk_import_id)
            return

        bulk_import.status = "processing"
        session.commit()

        # Download file from MinIO
        parts = bulk_import.file_path.split("/", 1)
        if len(parts) != 2:
            raise ValueError(f"Invalid file path format: {bulk_import.file_path}")
        content = download_file(parts[0], parts[1])

        # Detect format
        ext = bulk_import.filename.rsplit(".", 1)[-1].lower()

        rows = []
        if ext == "csv":
            try:
                text_content = content.decode("utf-8")
            except UnicodeDecodeError:
                text_content = content.decode("latin-1")
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
        elif ext in ("xlsx", "xls"):
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            wb.close()

        # Auto-detect column mapping
        def detect_column(row_dict, candidates):
            for key in row_dict.keys():
                if key:
                    key_lower = key.lower().strip()
                    for candidate in candidates:
                        if candidate in key_lower:
                            return key
            return None

        error_details = []

        for idx, row in enumerate(rows):
            try:
                # Detect name column
                name_key = detect_column(row, ["name", "nom"])
                name = row.get(name_key, "").strip() if name_key else ""

                if not name:
                    error_details.append(
                        {
                            "row": idx + 2,
                            "error": "Nom manquant",
                        }
                    )
                    bulk_import.error_count += 1
                    bulk_import.processed_count += 1
                    continue

                # Detect email column
                email_key = detect_column(row, ["email", "e-mail", "courriel", "mail"])
                email = row.get(email_key, "").strip() if email_key else None
                if email and not email:
                    email = None

                # Detect phone column
                phone_key = detect_column(row, ["phone", "telephone", "tel", "mobile"])
                phone = row.get(phone_key, "").strip() if phone_key else None
                if phone and not phone:
                    phone = None

                # Check deduplication by email
                if email:
                    existing = (
                        session.query(Candidate)
                        .filter(
                            Candidate.email == email,
                            Candidate.position_id == bulk_import.position_id,
                        )
                        .first()
                    )
                    if existing:
                        error_details.append(
                            {
                                "row": idx + 2,
                                "error": f"Email deja existant: {email}",
                            }
                        )
                        bulk_import.error_count += 1
                        bulk_import.processed_count += 1
                        continue

                # Create candidate
                candidate = Candidate(
                    tenant_id=bulk_import.tenant_id,
                    position_id=bulk_import.position_id,
                    name=name,
                    email=email,
                    phone=phone,
                    pipeline_status="new",
                )
                session.add(candidate)
                session.flush()

                # Create consent records
                for consent_type in ["data_processing", "call_recording"]:
                    consent = Consent(
                        candidate_id=candidate.id,
                        token=secrets.token_urlsafe(32),
                        type=consent_type,
                    )
                    session.add(consent)

                bulk_import.success_count += 1
                bulk_import.processed_count += 1
                session.commit()

            except Exception as e:
                logger.error("bulk_import_row_error", row=idx + 2, error=str(e))
                error_details.append(
                    {
                        "row": idx + 2,
                        "error": str(e),
                    }
                )
                bulk_import.error_count += 1
                bulk_import.processed_count += 1
                session.rollback()

        # Finalize
        bulk_import.status = "completed"
        bulk_import.completed_at = datetime.now(timezone.utc)
        if error_details:
            bulk_import.error_details = {"errors": error_details}
        session.commit()

        logger.info(
            "bulk_import_done",
            import_id=bulk_import_id,
            success=bulk_import.success_count,
            errors=bulk_import.error_count,
        )

    except Exception as e:
        session.rollback()
        logger.error("bulk_import_error", import_id=bulk_import_id, error=str(e))
        try:
            bulk_import = session.get(BulkImport, UUID(bulk_import_id))
            if bulk_import:
                bulk_import.status = "failed"
                bulk_import.completed_at = datetime.now(timezone.utc)
                bulk_import.error_details = {"error": str(e)}
                session.commit()
        except Exception:
            pass
        raise self.retry(exc=e, countdown=60)
    finally:
        session.close()
