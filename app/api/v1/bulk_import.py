import asyncio
import csv
import hashlib
import io
import json
import uuid as uuid_module
import zipfile
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.database import async_session, get_db
from app.core.dependencies import get_tenant_id, require_role
from app.models.bulk_import import BulkImport
from app.models.candidate import Candidate
from app.models.position import Position
from app.models.user import User
from app.schemas.bulk_import import (
    BulkImportBulkResponse,
    BulkImportDetail,
    BulkImportResponse,
    FileDecision,
    FilePreview,
    ImportConfirmRequest,
    ImportConfirmResponse,
    ImportPreviewResponse,
)
from app.services.storage import upload_file

router = APIRouter(prefix="/positions", tags=["Import"])
candidates_router = APIRouter(prefix="/candidates", tags=["Import"])
imports_router = APIRouter(prefix="/imports", tags=["Import"])
settings = get_settings()


@router.post(
    "/{position_id}/candidates/import-csv",
    response_model=BulkImportResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def import_csv(
    position_id: UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("admin", "recruiter")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Position).where(
            Position.id == position_id, Position.tenant_id == current_user.tenant_id
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Poste introuvable")

    if not file.filename:
        raise HTTPException(status_code=400, detail="Nom de fichier manquant")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(
            status_code=400,
            detail="Format de fichier non supporte. Utiliser .csv, .xlsx ou .xls",
        )

    content = await file.read()
    await file.seek(0)

    # Count rows
    total_count = 0
    if ext == "csv":
        try:
            text_content = content.decode("utf-8")
        except UnicodeDecodeError:
            text_content = content.decode("latin-1")
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)
        total_count = max(0, len(rows) - 1)  # Exclude header
    elif ext in ("xlsx", "xls"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        total_count = max(0, ws.max_row - 1)  # Exclude header
        wb.close()

    if total_count > 500:
        raise HTTPException(
            status_code=400, detail="Maximum 500 lignes autorisees par import"
        )

    # Upload file to MinIO
    await file.seek(0)
    file_path = await upload_file(
        file,
        "imports",
        f"{current_user.tenant_id}/{position_id}",
    )

    # Create BulkImport record
    bulk_import = BulkImport(
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        position_id=position_id,
        filename=file.filename,
        file_path=file_path,
        total_count=total_count,
        status="pending",
    )
    db.add(bulk_import)
    await db.flush()

    # Trigger worker
    from app.workers.bulk_import import process_csv_import

    process_csv_import.delay(str(bulk_import.id))

    return BulkImportResponse(
        import_id=str(bulk_import.id),
        filename=bulk_import.filename,
        total_count=bulk_import.total_count,
        status=bulk_import.status,
    )


@router.get("/{position_id}/imports/{import_id}/events")
async def import_events(
    position_id: UUID,
    import_id: UUID,
    request: Request,
    tenant_id: UUID = Depends(get_tenant_id),
):
    async def event_stream():
        last_processed = None
        last_status = None
        while True:
            if await request.is_disconnected():
                break
            async with async_session() as db:
                result = await db.execute(
                    select(BulkImport).where(
                        BulkImport.id == import_id,
                        BulkImport.position_id == position_id,
                        BulkImport.tenant_id == tenant_id,
                    )
                )
                bulk_import = result.scalar_one_or_none()
            if not bulk_import:
                yield f"event: error\ndata: {json.dumps({'detail': 'Import introuvable'})}\n\n"
                break
            processed_changed = bulk_import.processed_count != last_processed
            status_changed = bulk_import.status != last_status
            if processed_changed or status_changed:
                last_processed = bulk_import.processed_count
                last_status = bulk_import.status
                data = {
                    "processed": bulk_import.processed_count,
                    "total": bulk_import.total_count,
                    "success": bulk_import.success_count,
                    "errors": bulk_import.error_count,
                    "status": bulk_import.status,
                }
                yield f"event: progress\ndata: {json.dumps(data)}\n\n"
                if bulk_import.status in ("completed", "failed"):
                    yield f"event: done\ndata: {json.dumps({'status': bulk_import.status})}\n\n"
                    break
            await asyncio.sleep(1)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/{position_id}/imports", response_model=list[BulkImportDetail])
async def list_imports(
    position_id: UUID,
    tenant_id: UUID = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BulkImport)
        .where(BulkImport.position_id == position_id, BulkImport.tenant_id == tenant_id)
        .order_by(BulkImport.created_at.desc())
    )
    imports = result.scalars().all()

    return [
        BulkImportDetail(
            id=str(imp.id),
            filename=imp.filename,
            total_count=imp.total_count,
            processed_count=imp.processed_count,
            success_count=imp.success_count,
            error_count=imp.error_count,
            status=imp.status,
            error_details=imp.error_details,
            created_at=imp.created_at,
            completed_at=imp.completed_at,
        )
        for imp in imports
    ]


# ---------------------------------------------------------------------------
# Candidates router — Import massif de CVs (PDF/DOCX/ZIP)
# ---------------------------------------------------------------------------

ALLOWED_CV_EXTENSIONS = {"pdf", "docx"}
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_FILES_COUNT = 500


def _get_extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _extract_candidate_name(filename: str) -> str:
    """Extract candidate name from filename: 'Jean_Dupont_CV.pdf' -> 'Jean Dupont'"""
    name = filename.rsplit(".", 1)[0]  # Remove extension
    # Remove common suffixes
    for suffix in ["_CV", "_cv", "_Resume", "_resume", "-CV", "-cv", "(CV)", "[CV]"]:
        name = name.replace(suffix, "")
    # Replace separators with spaces
    name = name.replace("_", " ").replace("-", " ")
    # Clean up
    name = " ".join(name.split()).strip()
    return name or filename


async def _collect_entries(files: List[UploadFile]) -> list[tuple[str, bytes]]:
    """Expand uploads (including ZIPs) into (filename, content) pairs."""
    entries: list[tuple[str, bytes]] = []
    for upload in files:
        if not upload.filename:
            raise HTTPException(status_code=400, detail="Nom de fichier manquant")
        content = await upload.read()
        if len(content) > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"Fichier trop volumineux (max 10 Mo) : {upload.filename}",
            )
        ext = _get_extension(upload.filename)
        if ext == "zip":
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    for member in zf.infolist():
                        if member.is_dir():
                            continue
                        member_ext = _get_extension(member.filename)
                        if member_ext not in ALLOWED_CV_EXTENSIONS:
                            continue
                        member_content = zf.read(member.filename)
                        basename = member.filename.rsplit("/", 1)[-1]
                        entries.append((basename, member_content))
            except zipfile.BadZipFile:
                raise HTTPException(status_code=400, detail=f"ZIP invalide : {upload.filename}")
        elif ext in ALLOWED_CV_EXTENSIONS:
            entries.append((upload.filename, content))
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Format non supporte (PDF/DOCX/ZIP uniquement) : {upload.filename}",
            )
    return entries


# ---------------------------------------------------------------------------
# POST /candidates/import-bulk/preview
# ---------------------------------------------------------------------------

@candidates_router.post(
    "/import-bulk/preview",
    response_model=ImportPreviewResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def import_bulk_preview(
    files: List[UploadFile] = File(...),
    position_id: Optional[str] = Form(None),
    auto_score: bool = Form(True),
    current_user: User = Depends(require_role("admin", "recruiter")),
    db: AsyncSession = Depends(get_db),
):
    """Etape 1 : upload des fichiers, detection doublons, retourne la preview sans creer de candidats."""
    pos_uuid: Optional[UUID] = None
    if position_id:
        try:
            pos_uuid = UUID(position_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="position_id invalide")
        result = await db.execute(
            select(Position).where(
                Position.id == pos_uuid,
                Position.tenant_id == current_user.tenant_id,
            )
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="Poste introuvable")

    entries = await _collect_entries(files)

    if not entries:
        raise HTTPException(status_code=400, detail="Aucun fichier PDF/DOCX valide fourni")
    if len(entries) > MAX_FILES_COUNT:
        raise HTTPException(
            status_code=400,
            detail=f"Maximum {MAX_FILES_COUNT} fichiers par import (recu : {len(entries)})",
        )

    all_zip = all(_get_extension(u.filename or "") == "zip" for u in files)
    source_type = "zip" if all_zip else "files"

    # Upload each CV to MinIO
    from app.services.storage import s3_client, ensure_bucket

    bucket = "cvs"
    prefix = f"{current_user.tenant_id}/bulk"
    ensure_bucket(bucket)

    # Build preview list with duplicate detection
    file_previews: list[FilePreview] = []
    metadata_files: list[dict] = []

    for idx, (fname, fcontent) in enumerate(entries):
        ext = _get_extension(fname)
        candidate_name = _extract_candidate_name(fname)
        file_hash = hashlib.sha256(fcontent).hexdigest()
        size_bytes = len(fcontent)

        # Upload to MinIO
        key = f"{prefix}/{uuid_module.uuid4()}.{ext}"
        content_type = (
            "application/pdf"
            if ext == "pdf"
            else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        s3_client.put_object(Bucket=bucket, Key=key, Body=fcontent, ContentType=content_type)
        file_path = f"{bucket}/{key}"

        # Detect duplicate by SHA256 hash
        hash_result = await db.execute(
            select(Candidate).where(
                Candidate.tenant_id == current_user.tenant_id,
                Candidate.cv_parsed_data["file_hash"].astext == file_hash,
            )
        )
        existing_by_hash = hash_result.scalars().first()

        duplicate_info = None
        file_status = "new"

        if existing_by_hash:
            # Fetch position title if linked
            pos_title = None
            if existing_by_hash.position_id:
                pos_res = await db.execute(
                    select(Position).where(Position.id == existing_by_hash.position_id)
                )
                pos = pos_res.scalar_one_or_none()
                pos_title = pos.title if pos else None
            duplicate_info = {
                "candidate_id": str(existing_by_hash.id),
                "candidate_name": existing_by_hash.name,
                "position_title": pos_title,
                "match_type": "hash",
            }
            file_status = "duplicate"
        else:
            # Detect duplicate by name similarity (ILIKE)
            name_pattern = f"%{candidate_name}%"
            name_result = await db.execute(
                select(Candidate).where(
                    Candidate.tenant_id == current_user.tenant_id,
                    Candidate.name.ilike(name_pattern),
                )
            )
            existing_by_name = name_result.scalars().first()
            if existing_by_name:
                pos_title = None
                if existing_by_name.position_id:
                    pos_res = await db.execute(
                        select(Position).where(Position.id == existing_by_name.position_id)
                    )
                    pos = pos_res.scalar_one_or_none()
                    pos_title = pos.title if pos else None
                duplicate_info = {
                    "candidate_id": str(existing_by_name.id),
                    "candidate_name": existing_by_name.name,
                    "position_title": pos_title,
                    "match_type": "name",
                }
                file_status = "duplicate"

        preview = FilePreview(
            index=idx,
            filename=fname,
            candidate_name=candidate_name,
            size_bytes=size_bytes,
            status=file_status,
            error_message=None,
            duplicate_info=duplicate_info,
        )
        file_previews.append(preview)

        metadata_files.append({
            "index": idx,
            "filename": fname,
            "candidate_name": candidate_name,
            "file_path": file_path,
            "file_hash": file_hash,
            "size_bytes": size_bytes,
            "status": file_status,
            "duplicate_info": duplicate_info,
        })

    new_count = sum(1 for f in file_previews if f.status == "new")
    duplicate_count = sum(1 for f in file_previews if f.status == "duplicate")
    error_count = sum(1 for f in file_previews if f.status == "error")

    # Create BulkImport with status="preview"
    bulk_import = BulkImport(
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        position_id=pos_uuid,
        filename=f"bulk_{len(entries)}_cvs",
        file_path="bulk",
        total_count=len(entries),
        status="preview",
        source_type=source_type,
        import_metadata={
            "files": metadata_files,
            "auto_score": auto_score,
            "position_id": str(pos_uuid) if pos_uuid else None,
        },
    )
    db.add(bulk_import)
    await db.flush()
    await db.commit()

    return ImportPreviewResponse(
        import_id=str(bulk_import.id),
        total_count=len(entries),
        new_count=new_count,
        duplicate_count=duplicate_count,
        error_count=error_count,
        files=file_previews,
    )


# ---------------------------------------------------------------------------
# POST /candidates/import-bulk/{import_id}/confirm
# ---------------------------------------------------------------------------

@candidates_router.post(
    "/import-bulk/{import_id}/confirm",
    response_model=ImportConfirmResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def import_bulk_confirm(
    import_id: UUID,
    body: ImportConfirmRequest,
    current_user: User = Depends(require_role("admin", "recruiter")),
    db: AsyncSession = Depends(get_db),
):
    """Etape 2 : confirmer l'import avec les decisions par fichier (import/overwrite/skip)."""
    result = await db.execute(
        select(BulkImport).where(
            BulkImport.id == import_id,
            BulkImport.tenant_id == current_user.tenant_id,
            BulkImport.status == "preview",
        )
    )
    bulk_import = result.scalar_one_or_none()
    if not bulk_import:
        raise HTTPException(status_code=404, detail="Import introuvable ou deja confirme")

    metadata = bulk_import.import_metadata or {}
    metadata_files: list[dict] = metadata.get("files", [])
    auto_score: bool = metadata.get("auto_score", True)
    pos_id_str: Optional[str] = metadata.get("position_id")
    pos_uuid: Optional[UUID] = UUID(pos_id_str) if pos_id_str else None

    # Index decisions by file index
    decisions_map: dict[int, str] = {d.index: d.action for d in body.decisions}

    imported_count = 0
    overwritten_count = 0
    skipped_count = 0
    errors_count = 0
    celery_ids: list[str] = []

    from app.services.storage import s3_client

    for file_meta in metadata_files:
        idx = file_meta["index"]
        action = decisions_map.get(idx, "skip")
        file_status = file_meta.get("status", "new")

        # Errors are always skipped regardless of decision
        if file_status == "error":
            skipped_count += 1
            continue

        if action == "skip":
            skipped_count += 1
            continue

        file_path = file_meta["file_path"]  # e.g. "cvs/tenant_id/bulk/uuid.pdf"
        candidate_name = file_meta["candidate_name"]
        file_hash = file_meta["file_hash"]
        filename = file_meta["filename"]
        duplicate_info = file_meta.get("duplicate_info")

        try:
            if action == "overwrite" and duplicate_info:
                # Update existing candidate CV path and hash
                existing_id = UUID(duplicate_info["candidate_id"])
                cand_result = await db.execute(
                    select(Candidate).where(
                        Candidate.id == existing_id,
                        Candidate.tenant_id == current_user.tenant_id,
                    )
                )
                candidate = cand_result.scalar_one_or_none()
                if candidate:
                    candidate.cv_file_path = file_path
                    # Merge file_hash into cv_parsed_data
                    parsed = dict(candidate.cv_parsed_data or {})
                    parsed["file_hash"] = file_hash
                    candidate.cv_parsed_data = parsed
                    candidate.cv_score = None
                    candidate.cv_score_explanation = None
                    await db.flush()
                    celery_ids.append(str(candidate.id))
                    overwritten_count += 1
                else:
                    errors_count += 1
            else:
                # action == "import" — create new candidate (position_id nullable = vivier OK)
                candidate = Candidate(
                    tenant_id=current_user.tenant_id,
                    position_id=pos_uuid,  # may be None for vivier
                    name=candidate_name,
                    cv_file_path=file_path,
                    cv_parsed_data={"file_hash": file_hash, "original_filename": filename},
                    pipeline_status="new",
                )
                db.add(candidate)
                await db.flush()
                celery_ids.append(str(candidate.id))
                imported_count += 1
        except Exception:
            import structlog
            structlog.get_logger().warning(
                "import_confirm_file_error", index=idx, filename=filename
            )
            errors_count += 1

    # Update BulkImport record
    bulk_import.status = "processing"
    bulk_import.total_count = imported_count + overwritten_count
    bulk_import.import_metadata = {
        **metadata,
        "confirm_summary": {
            "imported": imported_count,
            "overwritten": overwritten_count,
            "skipped": skipped_count,
            "errors": errors_count,
        },
    }
    await db.commit()

    # Process CVs in parallel: try Celery first, fallback to inline async
    bid = str(import_id)
    if auto_score and celery_ids:
        celery_available = False
        try:
            from app.workers.cv_processing import process_cv

            for cid in celery_ids:
                process_cv.delay(cid, bulk_import_id=bid)
            celery_available = True
        except Exception:
            pass

        if not celery_available:
            # Fallback: process inline in background with concurrency limit
            import asyncio
            from starlette.concurrency import run_in_threadpool

            cids = list(celery_ids)

            async def _process_inline():
                import structlog as _structlog
                sem = asyncio.Semaphore(10)  # max 10 parallel

                async def _do_one(cid: str):
                    async with sem:
                        try:
                            from app.workers.cv_processing import process_cv as _pvc

                            await run_in_threadpool(_pvc, cid, None, bid)
                        except Exception as exc:
                            _structlog.get_logger().warning("inline_cv_error", candidate_id=cid, error=str(exc))

                await asyncio.gather(*[_do_one(c) for c in cids])

            asyncio.create_task(_process_inline())

    return ImportConfirmResponse(
        import_id=str(import_id),
        imported=imported_count,
        overwritten=overwritten_count,
        skipped=skipped_count,
        errors=errors_count,
        status="pending",
    )


# ---------------------------------------------------------------------------
# Legacy endpoint kept for backward compatibility
# ---------------------------------------------------------------------------

@candidates_router.post(
    "/import-bulk",
    response_model=BulkImportBulkResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def import_bulk_cvs(
    files: List[UploadFile] = File(...),
    position_id: Optional[str] = Form(None),
    auto_score: bool = Form(True),
    current_user: User = Depends(require_role("admin", "recruiter")),
    db: AsyncSession = Depends(get_db),
):
    """Import massif de CVs (PDF, DOCX, ZIP). Sans position_id = vivier de talents."""
    pos_uuid: Optional[UUID] = None
    if position_id:
        try:
            pos_uuid = UUID(position_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="position_id invalide")
        result = await db.execute(
            select(Position).where(
                Position.id == pos_uuid,
                Position.tenant_id == current_user.tenant_id,
            )
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="Poste introuvable")

    entries = await _collect_entries(files)

    if not entries:
        raise HTTPException(status_code=400, detail="Aucun fichier PDF/DOCX valide fourni")
    if len(entries) > MAX_FILES_COUNT:
        raise HTTPException(
            status_code=400,
            detail=f"Maximum {MAX_FILES_COUNT} fichiers par import (recu : {len(entries)})",
        )

    all_zip = all(_get_extension(u.filename or "") == "zip" for u in files)
    source_type = "zip" if all_zip else "files"

    from app.services.storage import s3_client, ensure_bucket

    bucket = "cvs"
    prefix = f"{current_user.tenant_id}/bulk"
    ensure_bucket(bucket)

    file_paths: list[str] = []
    filenames: list[str] = []
    for fname, fcontent in entries:
        ext = _get_extension(fname)
        key = f"{prefix}/{uuid_module.uuid4()}.{ext}"
        s3_client.put_object(
            Bucket=bucket,
            Key=key,
            Body=fcontent,
            ContentType=(
                "application/pdf"
                if ext == "pdf"
                else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            ),
        )
        file_paths.append(f"{bucket}/{key}")
        filenames.append(fname)

    bulk_import = BulkImport(
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        position_id=pos_uuid,
        filename=f"bulk_{len(entries)}_cvs",
        file_path="bulk",
        total_count=len(entries),
        status="pending",
        source_type=source_type,
        import_metadata={
            "file_paths": file_paths,
            "filenames": filenames,
            "auto_score": auto_score,
        },
    )
    db.add(bulk_import)
    await db.flush()
    await db.commit()

    try:
        from app.workers.bulk_import import process_bulk_cv_import

        process_bulk_cv_import.delay(str(bulk_import.id))
    except Exception:
        import structlog
        structlog.get_logger().warning("celery_unavailable", import_id=str(bulk_import.id))

    return BulkImportBulkResponse(
        import_id=str(bulk_import.id),
        total_count=bulk_import.total_count,
        status=bulk_import.status,
        source_type=source_type,
    )


# ---------------------------------------------------------------------------
# Imports router — SSE events sans position_id obligatoire
# ---------------------------------------------------------------------------


@imports_router.get("/recent")
async def list_recent_imports(
    current_user: User = Depends(require_role("admin", "recruiter")),
    db: AsyncSession = Depends(get_db),
):
    """List recent imports (last 20) for the current tenant, with metadata."""
    # Auto-cleanup: mark stuck imports (pending/processing with 0 progress for 30+ min) as failed
    from datetime import datetime, timedelta, timezone

    cutoff = datetime.now(timezone.utc) - timedelta(minutes=30)
    stuck_result = await db.execute(
        select(BulkImport).where(
            BulkImport.tenant_id == current_user.tenant_id,
            BulkImport.status.in_(["pending", "processing"]),
            BulkImport.processed_count == 0,
            BulkImport.created_at < cutoff,
        )
    )
    for stuck in stuck_result.scalars().all():
        stuck.status = "failed"
        stuck.completed_at = datetime.now(timezone.utc)
        stuck.error_details = {"error": "Import bloque (timeout 30min sans progression)"}
    await db.commit()

    recent_cutoff = datetime.now(timezone.utc) - timedelta(minutes=30)
    result = await db.execute(
        select(BulkImport)
        .where(
            BulkImport.tenant_id == current_user.tenant_id,
            # Only return: active imports OR imports from last 30 min
            or_(
                BulkImport.status.in_(["pending", "processing", "preview"]),
                BulkImport.created_at >= recent_cutoff,
            ),
        )
        .order_by(BulkImport.created_at.desc())
        .limit(20)
    )
    imports = result.scalars().all()

    return [
        {
            "id": str(imp.id),
            "filename": imp.filename,
            "total_count": imp.total_count,
            "processed_count": imp.processed_count,
            "success_count": imp.success_count,
            "error_count": imp.error_count,
            "status": imp.status,
            "source_type": imp.source_type,
            "created_at": imp.created_at.isoformat() if imp.created_at else None,
            "completed_at": imp.completed_at.isoformat() if imp.completed_at else None,
            "files": (imp.import_metadata or {}).get("files", []),
            "position_id": str(imp.position_id) if imp.position_id else None,
        }
        for imp in imports
    ]


@imports_router.get("/{import_id}/events")
async def import_events_global(
    import_id: UUID,
    request: Request,
    tenant_id: UUID = Depends(get_tenant_id),
):
    """SSE stream de progression pour un import (sans position_id requis)."""

    async def event_stream():
        last_processed = None
        last_status = None
        while True:
            if await request.is_disconnected():
                break
            async with async_session() as db:
                result = await db.execute(
                    select(BulkImport).where(
                        BulkImport.id == import_id,
                        BulkImport.tenant_id == tenant_id,
                    )
                )
                bulk_import = result.scalar_one_or_none()
            if not bulk_import:
                yield f"event: error\ndata: {json.dumps({'detail': 'Import introuvable'})}\n\n"
                break
            processed_changed = bulk_import.processed_count != last_processed
            status_changed = bulk_import.status != last_status
            if processed_changed or status_changed:
                last_processed = bulk_import.processed_count
                last_status = bulk_import.status
                data = {
                    "processed": bulk_import.processed_count,
                    "total": bulk_import.total_count,
                    "success": bulk_import.success_count,
                    "errors": bulk_import.error_count,
                    "status": bulk_import.status,
                }
                yield f"event: progress\ndata: {json.dumps(data)}\n\n"
                if bulk_import.status in ("completed", "failed"):
                    yield f"event: done\ndata: {json.dumps({'status': bulk_import.status})}\n\n"
                    break
            await asyncio.sleep(1)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )
