import json
import logging
import os
import uuid as uuid_mod
from datetime import datetime, timezone
from typing import Any, Dict
from uuid import UUID

from sqlalchemy import func, select, or_, and_, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.candidate import Candidate
from app.models.interview import Interview
from app.models.position import Position
from app.models.analysis import Analysis
from app.models.report import Report

logger = logging.getLogger(__name__)


async def handle_search_candidates(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    query = select(Candidate).where(Candidate.tenant_id == tenant_id)

    if params.get("position_id"):
        try:
            position_uuid = UUID(params["position_id"])
            query = query.where(Candidate.position_id == position_uuid)
        except (ValueError, AttributeError):
            pass

    if params.get("min_score") is not None:
        query = query.where(Candidate.cv_score >= params["min_score"])

    if params.get("max_score") is not None:
        query = query.where(Candidate.cv_score <= params["max_score"])

    if params.get("status"):
        query = query.where(Candidate.pipeline_status == params["status"])

    if params.get("search"):
        search_term = f"%{params['search']}%"
        query = query.where(
            or_(
                Candidate.name.ilike(search_term),
                Candidate.email.ilike(search_term)
            )
        )

    limit = min(params.get("limit", 20), 50)
    query = query.order_by(desc(Candidate.created_at)).limit(limit)

    result = await db.execute(query)
    candidates = result.scalars().all()

    data = []
    for c in candidates:
        data.append({
            "id": str(c.id),
            "name": c.name,
            "email": c.email,
            "position_id": str(c.position_id),
            "cv_score": c.cv_score,
            "pipeline_status": c.pipeline_status,
            "created_at": c.created_at.isoformat() if c.created_at else None
        })

    return json.dumps({
        "total": len(data),
        "candidates": data
    }, ensure_ascii=False, indent=2)


async def handle_list_positions(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    query = select(Position).where(Position.tenant_id == tenant_id)

    if params.get("status"):
        query = query.where(Position.status == params["status"])

    if params.get("search"):
        search_term = f"%{params['search']}%"
        query = query.where(
            or_(
                Position.title.ilike(search_term),
                Position.description.ilike(search_term)
            )
        )

    query = query.order_by(desc(Position.created_at)).limit(50)
    result = await db.execute(query)
    positions = result.scalars().all()

    data = []
    for p in positions:
        data.append({
            "id": str(p.id),
            "title": p.title,
            "seniority_level": p.seniority_level,
            "status": p.status,
            "created_at": p.created_at.isoformat() if p.created_at else None
        })

    return json.dumps({
        "total": len(data),
        "positions": data
    }, ensure_ascii=False, indent=2)


async def handle_get_position_details(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    try:
        position_uuid = UUID(params["position_id"])
    except (ValueError, KeyError):
        return json.dumps({"error": "position_id invalide"}, ensure_ascii=False)

    query = select(Position).where(
        and_(
            Position.id == position_uuid,
            Position.tenant_id == tenant_id
        )
    )
    result = await db.execute(query)
    position = result.scalar_one_or_none()

    if not position:
        return json.dumps({"error": "Poste non trouvé"}, ensure_ascii=False)

    count_query = select(func.count(Candidate.id)).where(
        and_(
            Candidate.position_id == position_uuid,
            Candidate.tenant_id == tenant_id
        )
    )
    count_result = await db.execute(count_query)
    candidate_count = count_result.scalar()

    data = {
        "id": str(position.id),
        "title": position.title,
        "description": position.description,
        "required_skills": position.required_skills,
        "seniority_level": position.seniority_level,
        "status": position.status,
        "created_at": position.created_at.isoformat() if position.created_at else None,
        "candidate_count": candidate_count
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_get_candidate_details(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    try:
        candidate_uuid = UUID(params["candidate_id"])
    except (ValueError, KeyError):
        return json.dumps({"error": "candidate_id invalide"}, ensure_ascii=False)

    query = select(Candidate).where(
        and_(
            Candidate.id == candidate_uuid,
            Candidate.tenant_id == tenant_id
        )
    )
    result = await db.execute(query)
    candidate = result.scalar_one_or_none()

    if not candidate:
        return json.dumps({"error": "Candidat non trouvé"}, ensure_ascii=False)

    position = None
    if candidate.position_id:
        pos_query = select(Position).where(Position.id == candidate.position_id)
        pos_result = await db.execute(pos_query)
        position = pos_result.scalar_one_or_none()

    interviews_query = select(Interview).where(
        and_(
            Interview.candidate_id == candidate_uuid,
            Interview.tenant_id == tenant_id
        )
    ).order_by(desc(Interview.created_at))
    interviews_result = await db.execute(interviews_query)
    interviews = interviews_result.scalars().all()

    interviews_data = []
    for interview in interviews:
        analysis = None
        if interview.id:
            analysis_query = select(Analysis).where(Analysis.interview_id == interview.id)
            analysis_result = await db.execute(analysis_query)
            analysis = analysis_result.scalar_one_or_none()

        interview_data = {
            "id": str(interview.id),
            "status": interview.status,
            "scheduled_at": interview.scheduled_at.isoformat() if interview.scheduled_at else None,
            "duration_seconds": interview.duration_seconds,
            "questions_asked": interview.questions_asked,
            "attempt_number": interview.attempt_number
        }

        if analysis:
            interview_data["analysis"] = {
                "skills_extracted": analysis.skills_extracted,
                "scores": analysis.scores,
                "score_explanations": analysis.score_explanations
            }

        interviews_data.append(interview_data)

    reports_query = select(Report).where(
        Report.candidate_id == candidate_uuid
    ).order_by(desc(Report.generated_at))
    reports_result = await db.execute(reports_query)
    reports = reports_result.scalars().all()

    reports_data = []
    for report in reports:
        reports_data.append({
            "id": str(report.id),
            "generated_at": report.generated_at.isoformat() if report.generated_at else None,
            "pdf_file_path": report.pdf_file_path
        })

    data = {
        "id": str(candidate.id),
        "name": candidate.name,
        "email": candidate.email,
        "phone": candidate.phone,
        "cv_score": candidate.cv_score,
        "cv_score_explanation": candidate.cv_score_explanation,
        "cv_parsed_data": candidate.cv_parsed_data,
        "pipeline_status": candidate.pipeline_status,
        "created_at": candidate.created_at.isoformat() if candidate.created_at else None,
        "position": {
            "id": str(position.id),
            "title": position.title,
            "seniority_level": position.seniority_level
        } if position else None,
        "interviews": interviews_data,
        "reports": reports_data
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_get_analytics_overview(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    total_candidates_query = select(func.count(Candidate.id)).where(
        Candidate.tenant_id == tenant_id
    )
    total_candidates_result = await db.execute(total_candidates_query)
    total_candidates = total_candidates_result.scalar()

    open_positions_query = select(func.count(Position.id)).where(
        and_(
            Position.tenant_id == tenant_id,
            Position.status == "active"
        )
    )
    open_positions_result = await db.execute(open_positions_query)
    open_positions = open_positions_result.scalar()

    avg_cv_score_query = select(func.avg(Candidate.cv_score)).where(
        and_(
            Candidate.tenant_id == tenant_id,
            Candidate.cv_score.is_not(None)
        )
    )
    avg_cv_score_result = await db.execute(avg_cv_score_query)
    avg_cv_score = avg_cv_score_result.scalar()

    completed_interviews_query = select(func.count(Interview.id)).where(
        and_(
            Interview.tenant_id == tenant_id,
            Interview.status == "completed"
        )
    )
    completed_interviews_result = await db.execute(completed_interviews_query)
    completed_interviews = completed_interviews_result.scalar()

    consent_given_query = select(func.count(Candidate.id)).where(
        and_(
            Candidate.tenant_id == tenant_id,
            Candidate.pipeline_status == "consent_given"
        )
    )
    consent_given_result = await db.execute(consent_given_query)
    consent_given = consent_given_result.scalar()

    conversion_rate = 0.0
    if total_candidates > 0:
        conversion_rate = (completed_interviews / total_candidates) * 100

    data = {
        "total_candidates": total_candidates,
        "active_positions": open_positions,
        "avg_cv_score": round(avg_cv_score, 2) if avg_cv_score else None,
        "completed_interviews": completed_interviews,
        "consent_given": consent_given,
        "conversion_rate_percent": round(conversion_rate, 2)
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_aggregate_scores(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    score_type = params["score_type"]
    position_id = params.get("position_id")
    position_uuid = None
    if position_id:
        try:
            position_uuid = UUID(position_id)
        except (ValueError, AttributeError):
            pass

    if score_type == "cv_score":
        query = select(
            func.avg(Candidate.cv_score).label("avg"),
            func.min(Candidate.cv_score).label("min"),
            func.max(Candidate.cv_score).label("max"),
            func.count(Candidate.id).label("count")
        ).where(
            and_(
                Candidate.tenant_id == tenant_id,
                Candidate.cv_score.is_not(None)
            )
        )

        if position_uuid:
            query = query.where(Candidate.position_id == position_uuid)

        result = await db.execute(query)
        row = result.one()

        data = {
            "score_type": score_type,
            "average": round(row.avg, 2) if row.avg else None,
            "min": round(row.min, 2) if row.min else None,
            "max": round(row.max, 2) if row.max else None,
            "count": row.count
        }

    else:
        query = select(Analysis.scores).select_from(Analysis).join(
            Interview, Interview.id == Analysis.interview_id
        ).join(
            Candidate, Candidate.id == Interview.candidate_id
        ).where(
            and_(
                Candidate.tenant_id == tenant_id,
                Analysis.scores.is_not(None)
            )
        )

        if position_uuid:
            query = query.where(Candidate.position_id == position_uuid)

        result = await db.execute(query)
        scores_list = [row[0] for row in result.fetchall()]

        values = []
        for scores_dict in scores_list:
            if isinstance(scores_dict, dict) and score_type in scores_dict:
                val = scores_dict[score_type]
                if val is not None:
                    values.append(float(val))

        if values:
            data = {
                "score_type": score_type,
                "average": round(sum(values) / len(values), 2),
                "min": round(min(values), 2),
                "max": round(max(values), 2),
                "count": len(values)
            }
        else:
            data = {
                "score_type": score_type,
                "average": None,
                "min": None,
                "max": None,
                "count": 0
            }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_get_pipeline_breakdown(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    query = select(
        Candidate.pipeline_status,
        func.count(Candidate.id).label("count")
    ).where(
        Candidate.tenant_id == tenant_id
    )

    if params.get("position_id"):
        try:
            position_uuid = UUID(params["position_id"])
            query = query.where(Candidate.position_id == position_uuid)
        except (ValueError, AttributeError):
            pass

    query = query.group_by(Candidate.pipeline_status)

    result = await db.execute(query)
    rows = result.fetchall()

    breakdown = {}
    total = 0
    for row in rows:
        s = row.pipeline_status or "unknown"
        count = row.count
        breakdown[s] = count
        total += count

    data = {
        "total": total,
        "breakdown": breakdown
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_export_data(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data_type = params.get("data_type", "candidates")
    position_id = params.get("position_id")
    status_filter = params.get("status")
    min_score = params.get("min_score")
    custom_filename = params.get("filename")

    position_uuid = None
    if position_id:
        try:
            position_uuid = UUID(position_id)
        except (ValueError, AttributeError):
            pass

    rows_data = []
    headers = []
    sheet_title = "Export"

    if data_type == "candidates":
        sheet_title = "Candidats"
        query = (
            select(Candidate, Position.title)
            .join(Position, Candidate.position_id == Position.id)
            .where(Candidate.tenant_id == tenant_id)
        )
        if position_uuid:
            query = query.where(Candidate.position_id == position_uuid)
        if status_filter:
            query = query.where(Candidate.pipeline_status == status_filter)
        if min_score is not None:
            query = query.where(Candidate.cv_score >= min_score)
        query = query.order_by(desc(Candidate.created_at)).limit(500)

        result = await db.execute(query)
        rows = result.all()

        headers = ["Nom", "Email", "Telephone", "Poste", "Score CV", "Statut", "Date"]
        for row in rows:
            c, pos_title = row
            rows_data.append([
                c.name,
                c.email or "",
                c.phone or "",
                pos_title,
                round(c.cv_score, 1) if c.cv_score is not None else "",
                c.pipeline_status,
                c.created_at.strftime("%d/%m/%Y") if c.created_at else "",
            ])

    elif data_type == "interviews":
        sheet_title = "Entretiens"
        query = (
            select(Interview, Candidate.name, Position.title)
            .join(Candidate, Interview.candidate_id == Candidate.id)
            .join(Position, Interview.position_id == Position.id)
            .where(Interview.tenant_id == tenant_id)
        )
        if position_uuid:
            query = query.where(Interview.position_id == position_uuid)
        if status_filter:
            query = query.where(Interview.status == status_filter)
        query = query.order_by(desc(Interview.created_at)).limit(500)

        result = await db.execute(query)
        rows = result.all()

        headers = ["Candidat", "Poste", "Statut", "Date planifiee", "Duree (s)", "Tentative"]
        for row in rows:
            itw, cand_name, pos_title = row
            rows_data.append([
                cand_name,
                pos_title,
                itw.status,
                itw.scheduled_at.strftime("%d/%m/%Y %H:%M") if itw.scheduled_at else "",
                itw.duration_seconds or "",
                itw.attempt_number,
            ])

    elif data_type == "positions":
        sheet_title = "Postes"
        query = select(Position).where(Position.tenant_id == tenant_id)
        if status_filter:
            query = query.where(Position.status == status_filter)
        query = query.order_by(desc(Position.created_at)).limit(200)

        result = await db.execute(query)
        positions = result.scalars().all()

        count_query = (
            select(Candidate.position_id, func.count(Candidate.id))
            .where(Candidate.tenant_id == tenant_id)
            .group_by(Candidate.position_id)
        )
        count_result = await db.execute(count_query)
        counts = {str(r[0]): r[1] for r in count_result.all()}

        headers = ["Titre", "Niveau", "Statut", "Candidats", "Date creation"]
        for p in positions:
            rows_data.append([
                p.title,
                p.seniority_level or "",
                p.status,
                counts.get(str(p.id), 0),
                p.created_at.strftime("%d/%m/%Y") if p.created_at else "",
            ])
    else:
        return json.dumps({"error": f"Type d'export inconnu: {data_type}"}, ensure_ascii=False)

    if not rows_data:
        return json.dumps({
            "error": "Aucune donnee trouvee pour les filtres specifies",
            "data_type": data_type,
            "filters": {"position_id": position_id, "status": status_filter, "min_score": min_score}
        }, ensure_ascii=False)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_data in rows_data:
            val = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")
    os.makedirs(exports_dir, exist_ok=True)

    file_id = uuid_mod.uuid4().hex[:12]
    if custom_filename:
        safe_name = "".join(c for c in custom_filename if c.isalnum() or c in "-_ ").strip()
        filename = f"{safe_name}_{file_id}.xlsx"
    else:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        filename = f"{data_type}_{timestamp}_{file_id}.xlsx"

    filepath = os.path.join(exports_dir, filename)
    wb.save(filepath)

    download_url = f"/api/v1/copilot/exports/{filename}"

    return json.dumps({
        "success": True,
        "filename": filename,
        "download_url": download_url,
        "row_count": len(rows_data),
        "columns": headers,
        "data_type": data_type,
    }, ensure_ascii=False)
