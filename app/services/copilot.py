"""
Service Copilot AIHM - Assistant IA pour interroger les données de recrutement.

Utilise Claude API avec tool_use pour répondre aux questions sur les candidats,
postes, scores, et analytics. Applique les LLM guardrails pour éviter les biais.
"""

import json
import logging
from typing import Any, Dict, List, Optional
from uuid import UUID

from sqlalchemy import func, select, or_, and_, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.candidate import Candidate
from app.models.interview import Interview
from app.models.position import Position
from app.models.analysis import Analysis
from app.models.report import Report

logger = logging.getLogger(__name__)

# System prompt avec guardrails LLM appliqués
COPILOT_SYSTEM_PROMPT = """Tu es l'assistant IA du système AIHM (AI Hiring Manager).

## Rôle
Tu aides les recruteurs à explorer et analyser les données de recrutement : candidats, entretiens, scores, postes.

## Guardrails stricts
- NE recommande JAMAIS d'embaucher ou rejeter un candidat
- NE déduis JAMAIS la personnalité, émotions, ou traits protégés (âge, genre, origine, religion, santé)
- Présente les données de façon factuelle et objective
- Si demandé un avis subjectif, rappelle que la décision appartient au recruteur
- Cite les scores et métriques sans interpréter la valeur humaine du candidat

## Comportement
- Réponds en français par défaut (sauf demande contraire)
- Utilise du markdown structuré (listes, tableaux) pour la lisibilité
- Explique brièvement quels outils tu utilises pour répondre
- Si les données sont insuffisantes, dis-le clairement
- Limite les résultats à 50 éléments max pour éviter la surcharge

## Outils disponibles
Tu as 8 outils pour interroger la base de données :
1. `search_candidates` : rechercher/filtrer des candidats
2. `list_positions` : lister les postes
3. `get_position_details` : détails d'un poste spécifique
4. `get_candidate_details` : fiche complète d'un candidat
5. `get_analytics_overview` : vue d'ensemble des KPIs
6. `aggregate_scores` : statistiques sur les scores
7. `get_pipeline_breakdown` : répartition des candidats par statut
8. `export_data` : exporter des données en fichier Excel (.xlsx) téléchargeable

Utilise ces outils pour répondre aux questions de façon précise et basée sur les données réelles.

## Export de données
Quand l'utilisateur demande un export, un téléchargement, un fichier Excel/CSV/XLS, utilise l'outil `export_data`.
Inclus TOUJOURS le lien de téléchargement dans ta réponse sous la forme : [Télécharger le fichier](URL)"""


# Définitions des outils Claude tool_use
COPILOT_TOOLS = [
    {
        "name": "search_candidates",
        "description": "Recherche et filtre les candidats selon divers critères (poste, score, statut, texte libre).",
        "input_schema": {
            "type": "object",
            "properties": {
                "position_id": {
                    "type": "string",
                    "description": "UUID du poste pour filtrer les candidats (optionnel)"
                },
                "min_score": {
                    "type": "number",
                    "description": "Score CV minimum (0-100, optionnel)"
                },
                "max_score": {
                    "type": "number",
                    "description": "Score CV maximum (0-100, optionnel)"
                },
                "status": {
                    "type": "string",
                    "description": "Statut pipeline : new, cv_uploaded, cv_analyzed, invited, consent_given, call_scheduled, call_in_progress, call_done, evaluated (optionnel)"
                },
                "search": {
                    "type": "string",
                    "description": "Texte libre pour rechercher dans nom, email (optionnel)"
                },
                "limit": {
                    "type": "integer",
                    "description": "Nombre max de résultats (défaut: 20, max: 50)"
                }
            }
        }
    },
    {
        "name": "list_positions",
        "description": "Liste tous les postes ouverts ou archivés.",
        "input_schema": {
            "type": "object",
            "properties": {
                "status": {
                    "type": "string",
                    "description": "Statut du poste : active, closed, draft (optionnel)"
                },
                "search": {
                    "type": "string",
                    "description": "Texte libre pour rechercher dans titre/description (optionnel)"
                }
            }
        }
    },
    {
        "name": "get_position_details",
        "description": "Récupère les détails complets d'un poste spécifique (compétences requises, seniority, etc.).",
        "input_schema": {
            "type": "object",
            "properties": {
                "position_id": {
                    "type": "string",
                    "description": "UUID du poste (requis)"
                }
            },
            "required": ["position_id"]
        }
    },
    {
        "name": "get_candidate_details",
        "description": "Fiche complète d'un candidat : infos perso, CV parsé, scores, entretiens, rapports.",
        "input_schema": {
            "type": "object",
            "properties": {
                "candidate_id": {
                    "type": "string",
                    "description": "UUID du candidat (requis)"
                }
            },
            "required": ["candidate_id"]
        }
    },
    {
        "name": "get_analytics_overview",
        "description": "Vue d'ensemble des KPIs : total candidats, postes ouverts, taux conversion, score moyen, etc.",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "aggregate_scores",
        "description": "Statistiques agrégées sur les scores (moyenne, min, max, distribution).",
        "input_schema": {
            "type": "object",
            "properties": {
                "score_type": {
                    "type": "string",
                    "description": "Type de score : cv_score, technical, experience, communication, global (requis)",
                    "enum": ["cv_score", "technical", "experience", "communication", "global"]
                },
                "position_id": {
                    "type": "string",
                    "description": "UUID du poste pour filtrer (optionnel)"
                }
            },
            "required": ["score_type"]
        }
    },
    {
        "name": "get_pipeline_breakdown",
        "description": "Répartition des candidats par statut pipeline (new, consent_pending, etc.).",
        "input_schema": {
            "type": "object",
            "properties": {
                "position_id": {
                    "type": "string",
                    "description": "UUID du poste pour filtrer (optionnel)"
                }
            }
        }
    },
    {
        "name": "export_data",
        "description": "Exporte des données en fichier Excel (.xlsx) téléchargeable. Utilise cet outil quand l'utilisateur demande un export, un téléchargement, un fichier Excel/CSV/XLS.",
        "input_schema": {
            "type": "object",
            "properties": {
                "data_type": {
                    "type": "string",
                    "description": "Type de données à exporter",
                    "enum": ["candidates", "interviews", "positions"]
                },
                "position_id": {
                    "type": "string",
                    "description": "UUID du poste pour filtrer (optionnel)"
                },
                "status": {
                    "type": "string",
                    "description": "Statut pour filtrer : pour candidates = pipeline_status (evaluated, cv_analyzed, etc.), pour positions = active/draft/closed, pour interviews = completed/scheduled/etc. (optionnel)"
                },
                "min_score": {
                    "type": "number",
                    "description": "Score CV minimum pour filtrer les candidats (optionnel)"
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Colonnes spécifiques à inclure (optionnel, toutes par défaut)"
                },
                "filename": {
                    "type": "string",
                    "description": "Nom du fichier sans extension (optionnel, auto-généré sinon)"
                }
            },
            "required": ["data_type"]
        }
    }
]


# ============================================================================
# HANDLERS DES OUTILS
# ============================================================================

async def handle_search_candidates(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Recherche candidats avec filtres multiples."""
    query = select(Candidate).where(Candidate.tenant_id == tenant_id)

    # Filtres
    if params.get("position_id"):
        try:
            position_uuid = UUID(params["position_id"])
            query = query.where(Candidate.position_id == position_uuid)
        except (ValueError, AttributeError):
            pass  # Invalid UUID, skip filter

    if params.get("min_score") is not None:
        query = query.where(Candidate.cv_score >= params["min_score"])

    if params.get("max_score") is not None:
        query = query.where(Candidate.cv_score <= params["max_score"])

    if params.get("status"):
        query = query.where(Candidate.pipeline_status == params["status"])

    if params.get("search"):
        search_term = f"%{params['search']}%"
        query = query.where(
            or_(
                Candidate.name.ilike(search_term),
                Candidate.email.ilike(search_term)
            )
        )

    # Tri et limite
    limit = min(params.get("limit", 20), 50)
    query = query.order_by(desc(Candidate.created_at)).limit(limit)

    result = await db.execute(query)
    candidates = result.scalars().all()

    # Formatter la réponse
    data = []
    for c in candidates:
        data.append({
            "id": str(c.id),
            "name": c.name,
            "email": c.email,
            "position_id": str(c.position_id),
            "cv_score": c.cv_score,
            "pipeline_status": c.pipeline_status,
            "created_at": c.created_at.isoformat() if c.created_at else None
        })

    return json.dumps({
        "total": len(data),
        "candidates": data
    }, ensure_ascii=False, indent=2)


async def handle_list_positions(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Liste les postes avec filtres."""
    query = select(Position).where(Position.tenant_id == tenant_id)

    if params.get("status"):
        query = query.where(Position.status == params["status"])

    if params.get("search"):
        search_term = f"%{params['search']}%"
        query = query.where(
            or_(
                Position.title.ilike(search_term),
                Position.description.ilike(search_term)
            )
        )

    query = query.order_by(desc(Position.created_at)).limit(50)
    result = await db.execute(query)
    positions = result.scalars().all()

    data = []
    for p in positions:
        data.append({
            "id": str(p.id),
            "title": p.title,
            "seniority_level": p.seniority_level,
            "status": p.status,
            "created_at": p.created_at.isoformat() if p.created_at else None
        })

    return json.dumps({
        "total": len(data),
        "positions": data
    }, ensure_ascii=False, indent=2)


async def handle_get_position_details(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Détails complets d'un poste."""
    try:
        position_uuid = UUID(params["position_id"])
    except (ValueError, KeyError):
        return json.dumps({"error": "position_id invalide"}, ensure_ascii=False)

    query = select(Position).where(
        and_(
            Position.id == position_uuid,
            Position.tenant_id == tenant_id
        )
    )
    result = await db.execute(query)
    position = result.scalar_one_or_none()

    if not position:
        return json.dumps({"error": "Poste non trouvé"}, ensure_ascii=False)

    # Compter les candidats pour ce poste
    count_query = select(func.count(Candidate.id)).where(
        and_(
            Candidate.position_id == position_uuid,
            Candidate.tenant_id == tenant_id
        )
    )
    count_result = await db.execute(count_query)
    candidate_count = count_result.scalar()

    data = {
        "id": str(position.id),
        "title": position.title,
        "description": position.description,
        "required_skills": position.required_skills,
        "seniority_level": position.seniority_level,
        "status": position.status,
        "created_at": position.created_at.isoformat() if position.created_at else None,
        "candidate_count": candidate_count
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_get_candidate_details(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Fiche complète candidat avec entretiens, analyses, rapports."""
    try:
        candidate_uuid = UUID(params["candidate_id"])
    except (ValueError, KeyError):
        return json.dumps({"error": "candidate_id invalide"}, ensure_ascii=False)

    # Candidat de base
    query = select(Candidate).where(
        and_(
            Candidate.id == candidate_uuid,
            Candidate.tenant_id == tenant_id
        )
    )
    result = await db.execute(query)
    candidate = result.scalar_one_or_none()

    if not candidate:
        return json.dumps({"error": "Candidat non trouvé"}, ensure_ascii=False)

    # Poste associé
    position = None
    if candidate.position_id:
        pos_query = select(Position).where(Position.id == candidate.position_id)
        pos_result = await db.execute(pos_query)
        position = pos_result.scalar_one_or_none()

    # Entretiens
    interviews_query = select(Interview).where(
        and_(
            Interview.candidate_id == candidate_uuid,
            Interview.tenant_id == tenant_id
        )
    ).order_by(desc(Interview.created_at))
    interviews_result = await db.execute(interviews_query)
    interviews = interviews_result.scalars().all()

    interviews_data = []
    for interview in interviews:
        # Analyse associée
        analysis = None
        if interview.id:
            analysis_query = select(Analysis).where(Analysis.interview_id == interview.id)
            analysis_result = await db.execute(analysis_query)
            analysis = analysis_result.scalar_one_or_none()

        interview_data = {
            "id": str(interview.id),
            "status": interview.status,
            "scheduled_at": interview.scheduled_at.isoformat() if interview.scheduled_at else None,
            "duration_seconds": interview.duration_seconds,
            "questions_asked": interview.questions_asked,
            "attempt_number": interview.attempt_number
        }

        if analysis:
            interview_data["analysis"] = {
                "skills_extracted": analysis.skills_extracted,
                "scores": analysis.scores,
                "score_explanations": analysis.score_explanations
            }

        interviews_data.append(interview_data)

    # Rapports
    reports_query = select(Report).where(
        Report.candidate_id == candidate_uuid
    ).order_by(desc(Report.generated_at))
    reports_result = await db.execute(reports_query)
    reports = reports_result.scalars().all()

    reports_data = []
    for report in reports:
        reports_data.append({
            "id": str(report.id),
            "generated_at": report.generated_at.isoformat() if report.generated_at else None,
            "pdf_file_path": report.pdf_file_path
        })

    # Construire la réponse complète
    data = {
        "id": str(candidate.id),
        "name": candidate.name,
        "email": candidate.email,
        "phone": candidate.phone,
        "cv_score": candidate.cv_score,
        "cv_score_explanation": candidate.cv_score_explanation,
        "cv_parsed_data": candidate.cv_parsed_data,
        "pipeline_status": candidate.pipeline_status,
        "created_at": candidate.created_at.isoformat() if candidate.created_at else None,
        "position": {
            "id": str(position.id),
            "title": position.title,
            "seniority_level": position.seniority_level
        } if position else None,
        "interviews": interviews_data,
        "reports": reports_data
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_get_analytics_overview(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Vue d'ensemble des KPIs."""
    # Total candidats
    total_candidates_query = select(func.count(Candidate.id)).where(
        Candidate.tenant_id == tenant_id
    )
    total_candidates_result = await db.execute(total_candidates_query)
    total_candidates = total_candidates_result.scalar()

    # Postes ouverts
    open_positions_query = select(func.count(Position.id)).where(
        and_(
            Position.tenant_id == tenant_id,
            Position.status == "active"
        )
    )
    open_positions_result = await db.execute(open_positions_query)
    open_positions = open_positions_result.scalar()

    # Score CV moyen
    avg_cv_score_query = select(func.avg(Candidate.cv_score)).where(
        and_(
            Candidate.tenant_id == tenant_id,
            Candidate.cv_score.is_not(None)
        )
    )
    avg_cv_score_result = await db.execute(avg_cv_score_query)
    avg_cv_score = avg_cv_score_result.scalar()

    # Entretiens complétés
    completed_interviews_query = select(func.count(Interview.id)).where(
        and_(
            Interview.tenant_id == tenant_id,
            Interview.status == "completed"
        )
    )
    completed_interviews_result = await db.execute(completed_interviews_query)
    completed_interviews = completed_interviews_result.scalar()

    # Candidats avec consentement
    consent_given_query = select(func.count(Candidate.id)).where(
        and_(
            Candidate.tenant_id == tenant_id,
            Candidate.pipeline_status == "consent_given"
        )
    )
    consent_given_result = await db.execute(consent_given_query)
    consent_given = consent_given_result.scalar()

    # Taux de conversion (completed / total)
    conversion_rate = 0.0
    if total_candidates > 0:
        conversion_rate = (completed_interviews / total_candidates) * 100

    data = {
        "total_candidates": total_candidates,
        "active_positions": open_positions,
        "avg_cv_score": round(avg_cv_score, 2) if avg_cv_score else None,
        "completed_interviews": completed_interviews,
        "consent_given": consent_given,
        "conversion_rate_percent": round(conversion_rate, 2)
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_aggregate_scores(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Statistiques agrégées sur les scores."""
    score_type = params["score_type"]
    position_id = params.get("position_id")
    position_uuid = None
    if position_id:
        try:
            position_uuid = UUID(position_id)
        except (ValueError, AttributeError):
            pass  # Invalid UUID, ignore filter

    if score_type == "cv_score":
        # Score CV : directement sur Candidate
        query = select(
            func.avg(Candidate.cv_score).label("avg"),
            func.min(Candidate.cv_score).label("min"),
            func.max(Candidate.cv_score).label("max"),
            func.count(Candidate.id).label("count")
        ).where(
            and_(
                Candidate.tenant_id == tenant_id,
                Candidate.cv_score.is_not(None)
            )
        )

        if position_uuid:
            query = query.where(Candidate.position_id == position_uuid)

        result = await db.execute(query)
        row = result.one()

        data = {
            "score_type": score_type,
            "average": round(row.avg, 2) if row.avg else None,
            "min": round(row.min, 2) if row.min else None,
            "max": round(row.max, 2) if row.max else None,
            "count": row.count
        }

    else:
        # Scores interview : depuis Analysis.scores JSONB
        # On doit faire un JOIN Candidate -> Interview -> Analysis
        query = select(Analysis.scores).select_from(Analysis).join(
            Interview, Interview.id == Analysis.interview_id
        ).join(
            Candidate, Candidate.id == Interview.candidate_id
        ).where(
            and_(
                Candidate.tenant_id == tenant_id,
                Analysis.scores.is_not(None)
            )
        )

        if position_uuid:
            query = query.where(Candidate.position_id == position_uuid)

        result = await db.execute(query)
        scores_list = [row[0] for row in result.fetchall()]

        # Extraire le score spécifique
        values = []
        for scores_dict in scores_list:
            if isinstance(scores_dict, dict) and score_type in scores_dict:
                val = scores_dict[score_type]
                if val is not None:
                    values.append(float(val))

        if values:
            data = {
                "score_type": score_type,
                "average": round(sum(values) / len(values), 2),
                "min": round(min(values), 2),
                "max": round(max(values), 2),
                "count": len(values)
            }
        else:
            data = {
                "score_type": score_type,
                "average": None,
                "min": None,
                "max": None,
                "count": 0
            }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_get_pipeline_breakdown(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Répartition des candidats par statut pipeline."""
    query = select(
        Candidate.pipeline_status,
        func.count(Candidate.id).label("count")
    ).where(
        Candidate.tenant_id == tenant_id
    )

    if params.get("position_id"):
        try:
            position_uuid = UUID(params["position_id"])
            query = query.where(Candidate.position_id == position_uuid)
        except (ValueError, AttributeError):
            pass  # Invalid UUID, skip filter

    query = query.group_by(Candidate.pipeline_status)

    result = await db.execute(query)
    rows = result.fetchall()

    breakdown = {}
    total = 0
    for row in rows:
        status = row.pipeline_status or "unknown"
        count = row.count
        breakdown[status] = count
        total += count

    data = {
        "total": total,
        "breakdown": breakdown
    }

    return json.dumps(data, ensure_ascii=False, indent=2)


async def handle_export_data(
    db: AsyncSession,
    tenant_id: UUID,
    params: Dict[str, Any]
) -> str:
    """Export data to an Excel file and return download URL."""
    import os
    import tempfile
    import uuid as uuid_mod
    from datetime import datetime, timezone
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data_type = params.get("data_type", "candidates")
    position_id = params.get("position_id")
    status_filter = params.get("status")
    min_score = params.get("min_score")
    custom_filename = params.get("filename")

    position_uuid = None
    if position_id:
        try:
            position_uuid = UUID(position_id)
        except (ValueError, AttributeError):
            pass

    rows_data = []
    headers = []
    sheet_title = "Export"

    if data_type == "candidates":
        sheet_title = "Candidats"
        query = (
            select(Candidate, Position.title)
            .join(Position, Candidate.position_id == Position.id)
            .where(Candidate.tenant_id == tenant_id)
        )
        if position_uuid:
            query = query.where(Candidate.position_id == position_uuid)
        if status_filter:
            query = query.where(Candidate.pipeline_status == status_filter)
        if min_score is not None:
            query = query.where(Candidate.cv_score >= min_score)
        query = query.order_by(desc(Candidate.created_at)).limit(500)

        result = await db.execute(query)
        rows = result.all()

        headers = ["Nom", "Email", "Telephone", "Poste", "Score CV", "Statut", "Date"]
        for row in rows:
            c, pos_title = row
            rows_data.append([
                c.name,
                c.email or "",
                c.phone or "",
                pos_title,
                round(c.cv_score, 1) if c.cv_score is not None else "",
                c.pipeline_status,
                c.created_at.strftime("%d/%m/%Y") if c.created_at else "",
            ])

    elif data_type == "interviews":
        sheet_title = "Entretiens"
        query = (
            select(Interview, Candidate.name, Position.title)
            .join(Candidate, Interview.candidate_id == Candidate.id)
            .join(Position, Interview.position_id == Position.id)
            .where(Interview.tenant_id == tenant_id)
        )
        if position_uuid:
            query = query.where(Interview.position_id == position_uuid)
        if status_filter:
            query = query.where(Interview.status == status_filter)
        query = query.order_by(desc(Interview.created_at)).limit(500)

        result = await db.execute(query)
        rows = result.all()

        headers = ["Candidat", "Poste", "Statut", "Date planifiee", "Duree (s)", "Tentative"]
        for row in rows:
            itw, cand_name, pos_title = row
            rows_data.append([
                cand_name,
                pos_title,
                itw.status,
                itw.scheduled_at.strftime("%d/%m/%Y %H:%M") if itw.scheduled_at else "",
                itw.duration_seconds or "",
                itw.attempt_number,
            ])

    elif data_type == "positions":
        sheet_title = "Postes"
        query = select(Position).where(Position.tenant_id == tenant_id)
        if status_filter:
            query = query.where(Position.status == status_filter)
        query = query.order_by(desc(Position.created_at)).limit(200)

        result = await db.execute(query)
        positions = result.scalars().all()

        # Count candidates per position
        count_query = (
            select(Candidate.position_id, func.count(Candidate.id))
            .where(Candidate.tenant_id == tenant_id)
            .group_by(Candidate.position_id)
        )
        count_result = await db.execute(count_query)
        counts = {str(r[0]): r[1] for r in count_result.all()}

        headers = ["Titre", "Niveau", "Statut", "Candidats", "Date creation"]
        for p in positions:
            rows_data.append([
                p.title,
                p.seniority_level or "",
                p.status,
                counts.get(str(p.id), 0),
                p.created_at.strftime("%d/%m/%Y") if p.created_at else "",
            ])
    else:
        return json.dumps({"error": f"Type d'export inconnu: {data_type}"}, ensure_ascii=False)

    if not rows_data:
        return json.dumps({
            "error": "Aucune donnee trouvee pour les filtres specifies",
            "data_type": data_type,
            "filters": {"position_id": position_id, "status": status_filter, "min_score": min_score}
        }, ensure_ascii=False)

    # Generate Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(rows_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_data in rows_data:
            val = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # Save to exports directory
    exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")
    os.makedirs(exports_dir, exist_ok=True)

    file_id = uuid_mod.uuid4().hex[:12]
    if custom_filename:
        safe_name = "".join(c for c in custom_filename if c.isalnum() or c in "-_ ").strip()
        filename = f"{safe_name}_{file_id}.xlsx"
    else:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        filename = f"{data_type}_{timestamp}_{file_id}.xlsx"

    filepath = os.path.join(exports_dir, filename)
    wb.save(filepath)

    download_url = f"/api/v1/copilot/exports/{filename}"

    return json.dumps({
        "success": True,
        "filename": filename,
        "download_url": download_url,
        "row_count": len(rows_data),
        "columns": headers,
        "data_type": data_type,
    }, ensure_ascii=False)


# ============================================================================
# DISPATCHER
# ============================================================================

async def execute_tool(
    tool_name: str,
    tool_input: Dict[str, Any],
    db: AsyncSession,
    tenant_id: UUID
) -> str:
    """
    Exécute un outil Copilot et retourne le résultat en JSON string.

    Args:
        tool_name: Nom de l'outil à exécuter
        tool_input: Paramètres d'entrée de l'outil
        db: Session DB async
        tenant_id: ID du tenant (UUID) pour filtrage multi-tenant

    Returns:
        JSON string avec le résultat
    """
    logger.info(f"Copilot tool execution: {tool_name} with params {tool_input}")

    try:
        if tool_name == "search_candidates":
            return await handle_search_candidates(db, tenant_id, tool_input)

        elif tool_name == "list_positions":
            return await handle_list_positions(db, tenant_id, tool_input)

        elif tool_name == "get_position_details":
            return await handle_get_position_details(db, tenant_id, tool_input)

        elif tool_name == "get_candidate_details":
            return await handle_get_candidate_details(db, tenant_id, tool_input)

        elif tool_name == "get_analytics_overview":
            return await handle_get_analytics_overview(db, tenant_id, tool_input)

        elif tool_name == "aggregate_scores":
            return await handle_aggregate_scores(db, tenant_id, tool_input)

        elif tool_name == "get_pipeline_breakdown":
            return await handle_get_pipeline_breakdown(db, tenant_id, tool_input)

        elif tool_name == "export_data":
            return await handle_export_data(db, tenant_id, tool_input)

        else:
            return json.dumps({
                "error": f"Outil inconnu : {tool_name}"
            }, ensure_ascii=False)

    except Exception as e:
        logger.error(f"Error executing tool {tool_name}: {e}", exc_info=True)
        return json.dumps({
            "error": f"Erreur lors de l'exécution de l'outil : {str(e)}"
        }, ensure_ascii=False)
